"""Import an Impact List from FEMA's damage inventory workbook.

Every applicant on a declaration receives the same template from their state
recipient, so this reads the form rather than any one jurisdiction's file. The
header block (disaster number, applicant name, FIPS) sits in the first rows; the
site rows begin under a header row whose first data column is "Category".

Contact names, phone numbers, and email addresses in the header block are read but
never stored on the Scenario -- the engine has no field for them, by design. What
gets kept is the applicant's role, not the individual.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from .models import Applicant, Disaster, Scenario, Site
from .rules import DEFAULT_RULES, LABOR_TYPES, PRIMARY_CAUSES

_HEADER_KEYS = {
    "disaster number:": "disaster_number",
    "applicant name:": "applicant_name",
    "applicant fips:": "applicant_fips",
}

_COLUMNS = [
    "category", "name", "address1", "address2", "city", "state", "zip",
    "latitude", "longitude", "damage_description", "primary_cause",
    "approx_cost", "percent_complete", "labor_type", "prior_pa_grant", "priority",
]


def _clean(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def _percent(v: Any) -> float:
    n = _number(v)
    if n is None:
        return 0.0
    # The template accepts either 0.9 or 90 for ninety percent.
    return round(n / 100.0 if n > 1.0 else n, 4)


def _normalize_cause(raw: str) -> str:
    r = raw.strip().lower()
    for cause in PRIMARY_CAUSES:
        if cause.lower() == r:
            return cause
    return raw.strip()


def _fix_longitude(lon: float | None, state: str) -> float | None:
    """Western-hemisphere longitudes are negative; the template is often keyed
    without the sign. Correct it rather than silently carrying a bad coordinate."""
    if lon is None:
        return None
    if lon > 0 and state.upper() in {
        "WA", "OR", "CA", "ID", "MT", "WY", "NV", "UT", "AZ", "NM", "CO", "AK", "HI",
    }:
        return -lon
    return lon


def read_damage_inventory(path: str | Path) -> tuple[dict[str, str], list[Site]]:
    """Return (header fields, sites) from one damage inventory workbook."""
    import openpyxl  # imported lazily so the engine works without openpyxl installed

    wb = openpyxl.load_workbook(Path(path), data_only=True)
    ws = wb["Damage Inventory"] if "Damage Inventory" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    header: dict[str, str] = {}
    data_start = None
    col_offset = 0

    for i, row in enumerate(rows):
        cells = [_clean(c) for c in row]
        for j, cell in enumerate(cells):
            key = _HEADER_KEYS.get(cell.lower())
            if key:
                value = next((c for c in cells[j + 1:] if c), "")
                header[key] = value
            if cell.lower() == "category" and data_start is None:
                data_start = i + 1
                col_offset = j
    if data_start is None:
        return header, []

    sites: list[Site] = []
    for row in rows[data_start:]:
        vals = list(row[col_offset:col_offset + len(_COLUMNS)])
        vals += [None] * (len(_COLUMNS) - len(vals))
        rec = dict(zip(_COLUMNS, vals))

        category = _clean(rec["category"]).upper()
        name = _clean(rec["name"])
        if not name:
            continue
        # The template's footer repeats the labor-type legend; skip it.
        if name.lower().startswith("labor key"):
            continue

        state = _clean(rec["state"])
        address = " ".join(x for x in (_clean(rec["address1"]), _clean(rec["address2"])) if x)
        labor = _clean(rec["labor_type"]).upper()

        sites.append(Site(
            category=category or "B",
            name=name,
            address=address,
            city=_clean(rec["city"]),
            state=state,
            zip_code=_clean(rec["zip"]),
            latitude=_number(rec["latitude"]),
            longitude=_fix_longitude(_number(rec["longitude"]), state),
            damage_description=_clean(rec["damage_description"]),
            primary_cause=_normalize_cause(_clean(rec["primary_cause"])),
            approx_cost=_number(rec["approx_cost"]) or 0.0,
            percent_complete=_percent(rec["percent_complete"]),
            labor_type=labor if labor in LABOR_TYPES else "FA",
            prior_pa_grant=(_clean(rec["prior_pa_grant"]).upper() or "U")[:1],
            priority=_clean(rec["priority"]),
        ))
    return header, sites


def scenario_from_inventories(
    paths: list[str | Path],
    title: str = "",
    disaster_name: str = "",
    **disaster_kwargs,
) -> Scenario:
    """Build a Scenario from one or more damage inventory workbooks.

    Applicants routinely receive the inventory split across several files (one per
    department, or one per category block), so this merges them and de-duplicates
    on category + site name.
    """
    header: dict[str, str] = {}
    sites: list[Site] = []
    seen: set[tuple[str, str]] = set()

    for p in paths:
        h, s = read_damage_inventory(p)
        header = {**h, **{k: v for k, v in header.items() if v}}
        for site in s:
            key = (site.category, site.name.lower())
            if key in seen:
                continue
            seen.add(key)
            sites.append(site)

    applicant_name = header.get("applicant_name", "")
    scenario = Scenario(
        title=title or f"{header.get('disaster_number', 'Disaster')} — {applicant_name}",
        applicant=Applicant(
            name=applicant_name,
            fips=header.get("applicant_fips", ""),
            entity_type="Local Government",
        ),
        disaster=Disaster(
            number=header.get("disaster_number", ""),
            name=disaster_name,
            **disaster_kwargs,
        ),
        rules=DEFAULT_RULES,
        sites=sites,
    )
    return scenario
